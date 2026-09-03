import os
import io
import csv
import zipfile
import xml.etree.ElementTree as ET
from typing import List, Dict, Any, Tuple, Optional


def parse_csv_data(file_bytes: bytes, filename: str = "data.csv") -> Tuple[List[str], List[Dict[str, Any]], Dict[str, Any]]:
    """Parse CSV data into headers and list of row dicts."""
    # Attempt decoding with UTF-8, then fallback to latin-1 / cp1252
    text = ""
    for enc in ["utf-8-sig", "utf-8", "latin-1", "cp1252"]:
        try:
            text = file_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue

    if not text:
        text = file_bytes.decode("utf-8", errors="replace")

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)

    if not rows:
        return [], [], {"error": "Empty CSV file"}

    # Find the header row (first non-empty row)
    header_idx = 0
    while header_idx < len(rows) and not any(cell.strip() for cell in rows[header_idx]):
        header_idx += 1

    if header_idx >= len(rows):
        return [], [], {"error": "No header or data rows found in CSV"}

    raw_headers = [h.strip() for h in rows[header_idx]]
    # Ensure unique headers
    headers = []
    seen = {}
    for idx, h in enumerate(raw_headers):
        clean_h = h if h else f"Column_{idx + 1}"
        if clean_h in seen:
            seen[clean_h] += 1
            headers.append(f"{clean_h}_{seen[clean_h]}")
        else:
            seen[clean_h] = 0
            headers.append(clean_h)

    data_rows = []
    for row_idx in range(header_idx + 1, len(rows)):
        raw_row = rows[row_idx]
        if not any(cell.strip() for cell in raw_row):
            continue  # Skip completely blank lines

        row_dict = {}
        for col_idx, header in enumerate(headers):
            val = raw_row[col_idx].strip() if col_idx < len(raw_row) else ""
            row_dict[header] = val
        data_rows.append(row_dict)

    metadata = {
        "sheet_name": "CSV",
        "total_rows": len(data_rows),
        "embedded_images": 0
    }
    return headers, data_rows, metadata


def extract_embedded_images_from_xlsx(file_path: str, extract_dir: str) -> Dict[int, str]:
    """
    Extract embedded images from an XLSX workbook zip structure and map by row index if possible.
    Returns a dict of row_index (1-based) -> extracted_image_path.
    """
    row_to_image = {}
    os.makedirs(extract_dir, exist_ok=True)

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        # openpyxl keeps images in ws._images
        if hasattr(ws, "_images") and ws._images:
            for idx, img in enumerate(ws._images):
                try:
                    # Get anchor row
                    row = None
                    if hasattr(img, "anchor"):
                        anchor = img.anchor
                        if hasattr(anchor, "_from") and hasattr(anchor._from, "row"):
                            row = anchor._from.row + 1  # 1-indexed
                        elif hasattr(anchor, "row"):
                            row = anchor.row

                    if row is None:
                        row = idx + 2  # default assume sequential starting after header

                    img_filename = f"embedded_row_{row}_{idx + 1}.png"
                    save_path = os.path.join(extract_dir, img_filename)

                    # Extract image data
                    image_data = img._data()
                    with open(save_path, "wb") as f:
                        f.write(image_data)

                    row_to_image[row] = save_path
                except Exception as img_err:
                    print(f"[IMPORTER] Could not extract image {idx}: {img_err}")
    except Exception as e:
        print(f"[IMPORTER] Openpyxl image extraction fallback: {e}")

    # Fallback to direct ZIP inspection if openpyxl found no images
    if not row_to_image:
        try:
            with zipfile.ZipFile(file_path, 'r') as z:
                media_files = [f for f in z.namelist() if f.startswith('xl/media/')]
                for idx, media in enumerate(media_files):
                    img_data = z.read(media)
                    ext = media.split('.')[-1]
                    target_filename = f"embedded_img_{idx + 1}.{ext}"
                    save_path = os.path.join(extract_dir, target_filename)
                    with open(save_path, "wb") as f:
                        f.write(img_data)
                    # Associate sequentially starting from row 2
                    row_to_image[idx + 2] = save_path
        except Exception as zip_err:
            print(f"[IMPORTER] Zip fallback error: {zip_err}")

    return row_to_image


def parse_xlsx_data(file_path: str, sheet_name: Optional[str] = None, extract_embedded: bool = True, extract_dir: Optional[str] = None) -> Tuple[List[str], List[Dict[str, Any]], Dict[str, Any]]:
    """Parse XLSX worksheet and return headers, rows, and metadata."""
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet_names = wb.sheetnames

    target_sheet = sheet_name if sheet_name and sheet_name in sheet_names else sheet_names[0]
    ws = wb[target_sheet]

    # Find header row
    header_row_idx = 1
    max_r = min(ws.max_row, 10)
    for r in range(1, max_r + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None and str(v).strip() for v in row_vals):
            header_row_idx = r
            break

    # Read headers
    raw_headers = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(header_row_idx, c).value
        clean = str(val).strip() if val is not None else f"Column_{c}"
        raw_headers.append(clean)

    # Unique headers
    headers = []
    seen = {}
    for idx, h in enumerate(raw_headers):
        if h in seen:
            seen[h] += 1
            headers.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            headers.append(h)

    # Extract embedded images if requested
    embedded_images = {}
    if extract_embedded and extract_dir:
        embedded_images = extract_embedded_images_from_xlsx(file_path, extract_dir)

    # Read data rows
    data_rows = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
        # Skip completely empty rows
        if not any(v is not None and str(v).strip() for v in row_vals):
            continue

        row_dict = {}
        for col_idx, header in enumerate(headers):
            val = row_vals[col_idx] if col_idx < len(row_vals) else None
            # Normalize cell types
            if val is None:
                str_val = ""
            elif isinstance(val, float) and val.is_integer():
                str_val = str(int(val))
            else:
                str_val = str(val).strip()
            row_dict[header] = str_val

        # Attach embedded image reference if found for this row
        if r in embedded_images:
            row_dict["_embedded_photo_path"] = embedded_images[r]

        data_rows.append(row_dict)

    metadata = {
        "sheets": sheet_names,
        "selected_sheet": target_sheet,
        "total_rows": len(data_rows),
        "embedded_images_count": len(embedded_images)
    }

    return headers, data_rows, metadata
